from collections import Counter
import csv
import io
import html
import json
import logging
import os
import re
from contextlib import asynccontextmanager
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, Query, File, Form
from fastapi.responses import FileResponse
from openpyxl import load_workbook
import aiosqlite

from database import (
    init_db,
    get_trader,
    get_traders,
    create_trader,
    update_trader,
    delete_trader,
    get_all_settings,
    get_setting,
    set_setting,
    import_traders,
    get_traders_stats,
    log_outreach,
    get_outreach_logs,
    DB_PATH,
)


def escape_html(text: str) -> str:
    """Escape HTML special characters to prevent XSS."""
    if text is None or text == "":
        return ""
    return html.escape(str(text))


STOPWORDS = frozenset({
    'the', 'a', 'an', 'is', 'are', 'was', 'were', 'be', 'been', 'being',
    'have', 'has', 'had', 'do', 'does', 'did', 'will', 'would', 'could',
    'should', 'may', 'might', 'shall', 'can', 'to', 'of', 'in', 'for',
    'on', 'with', 'at', 'by', 'from', 'as', 'into', 'through', 'during',
    'before', 'after', 'above', 'below', 'between', 'out', 'off', 'over',
    'under', 'again', 'further', 'then', 'once', 'and', 'but', 'or', 'nor',
    'not', 'no', 'so', 'if', 'too', 'very', 'just', 'about', 'also',
    'this', 'that', 'these', 'those', 'it', 'its', 'my', 'your', 'his',
    'her', 'our', 'their', 'what', 'which', 'who', 'whom', 'where', 'when',
    'why', 'how', 'all', 'each', 'every', 'both', 'few', 'more', 'most',
    'other', 'some', 'such', 'only', 'own', 'same', 'here', 'there',
})


def analyze_trader_keywords(raw_text: str) -> dict:
    """Extract keywords from raw text without external APIs.
    
    Uses word frequency analysis with abbreviation-aware sentence splitting.
    """
    if not raw_text or not raw_text.strip():
        return {"keywords": [], "summary": "No text provided for analysis"}
    
    # Word frequency analysis
    words = raw_text.lower().split()
    word_freq = Counter(
        clean for word in words
        if (clean := ''.join(c for c in word if c.isalnum()))
        and len(clean) > 2
        and clean not in STOPWORDS
    )
    
    # Top keywords by frequency
    keywords = [w for w, _ in word_freq.most_common(10)]
    
    # Abbreviation-aware sentence splitting
    sentences = re.split(r'(?<=[.!?])\s+(?=[A-Z])', raw_text)
    sentences = [s.strip() for s in sentences if len(s.strip()) > 10]
    
    if sentences:
        summary = sentences[0]
        if len(sentences) > 1:
            summary += ' ' + sentences[1]
    else:
        summary = raw_text[:200]
    
    if len(summary) > 200:
        summary = summary[:197] + '...'
    
    return {"keywords": keywords, "summary": summary}

load_dotenv()

PORT = int(os.getenv("PORT", "8000"))

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(name)s] %(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

INDEX_PATH = Path(__file__).parent / "index.html"
CSS_PATH = Path(__file__).parent / "design-system.css"


@asynccontextmanager
async def lifespan(app: FastAPI):
    # Startup: initialize database and seed sample data if empty
    await init_db()
    async with aiosqlite.connect(DB_PATH) as conn:
        cursor = await conn.execute("SELECT COUNT(*) FROM traders")
        count = (await cursor.fetchone())[0]
        if count == 0:
            await _seed_sample_data(conn)
    logger.info("Database initialized")
    yield


async def _seed_sample_data(conn):
    """Insert sample traders if the database is empty."""
    samples = [
        {
            "trader_name": "CryptoKing",
            "platform": "eToro",
            "profile_url": "https://etoro.com/cryptoking",
            "location": "Singapore",
            "language": "English",
            "strategy_type": "DeFi Yield",
            "twitter": "@cryptoking",
            "telegram": "@cryptoking",
            "followers": 15000,
            "copiers": 3400,
            "assets_under_management": 25000000.0,
            "monthly_return": 12.5,
            "yearly_return": 145.0,
            "maximum_drawdown": 8.5,
            "risk_score": 7,
            "audience_strength": 9,
            "trading_consistency": 8,
            "communication_quality": 6,
            "crypto_knowledge": 10,
            "likelihood_to_join": 8,
            "brand_value": 7,
            "fit_score": 81,
            "creator_score": 75,
            "research_notes": "Strong DeFi presence. High AUM. Might be selective due to existing following.",
            "tags": "DeFi, High Volume, KOL",
            "status": "Contacted",
            "pipeline_stage": "Outreach",
            "first_contact_date": "2025-06-01T10:00:00Z",
            "last_contact_date": "2025-06-03T14:30:00Z",
            "date_added": "2025-05-28T08:00:00Z",
            "date_updated": "2025-06-03T14:30:00Z",
            "outreach_attempts": 2,
        },
        {
            "trader_name": "FXQuant",
            "platform": "MQL5",
            "profile_url": "https://mql5.com/users/fxquant",
            "location": "London, UK",
            "language": "English",
            "strategy_type": "Algorithmic Trading",
            "telegram": "@fxquant_signals",
            "youtube": "https://youtube.com/@fxquant",
            "followers": 8200,
            "copiers": 1200,
            "assets_under_management": 8500000.0,
            "monthly_return": 4.2,
            "yearly_return": 52.0,
            "maximum_drawdown": 12.0,
            "risk_score": 6,
            "audience_strength": 7,
            "trading_consistency": 9,
            "communication_quality": 8,
            "crypto_knowledge": 3,
            "likelihood_to_join": 6,
            "brand_value": 6,
            "fit_score": 65,
            "creator_score": 60,
            "research_notes": "Experienced algo trader. Traditional forex background. Will need education on DeFi.",
            "tags": "Algorithmic, Traditional Finance",
            "status": "Interested",
            "pipeline_stage": "Interview",
            "date_added": "2025-05-20T11:00:00Z",
            "date_contacted": "2025-05-25T09:00:00Z",
            "date_replied": "2025-05-26T16:00:00Z",
            "date_meeting_scheduled": "2025-06-05T15:00:00Z",
            "last_contact_date": "2025-06-04T11:00:00Z",
            "date_updated": "2025-06-04T11:00:00Z",
            "outreach_attempts": 3,
        },
        {
            "trader_name": "DeFiDegen",
            "platform": "Manual",
            "profile_url": "https://twitter.com/defidegen",
            "location": "Crypto Valley, Switzerland",
            "language": "English",
            "strategy_type": "Liquidity Pool",
            "twitter": "@defidegen",
            "discord": "DeFiDegen#1234",
            "followers": 4500,
            "copiers": 0,
            "assets_under_management": 1200000.0,
            "monthly_return": 25.0,
            "yearly_return": 320.0,
            "maximum_drawdown": 18.0,
            "risk_score": 9,
            "audience_strength": 6,
            "trading_consistency": 6,
            "communication_quality": 7,
            "crypto_knowledge": 10,
            "likelihood_to_join": 9,
            "brand_value": 5,
            "fit_score": 72,
            "creator_score": 70,
            "research_notes": "Extremely high returns but also high drawdowns. DeFi native. Might be good yield strategies.",
            "tags": "High Risk, DeFi Native, Yield Farmer",
            "status": "Meeting Scheduled",
            "pipeline_stage": "Interview",
            "date_added": "2025-06-02T14:00:00Z",
            "date_researched": "2025-06-02T16:00:00Z",
            "date_contacted": "2025-06-03T10:00:00Z",
            "last_contact_date": "2025-06-04T09:00:00Z",
            "date_updated": "2025-06-04T09:00:00Z",
            "outreach_attempts": 2,
        },
        {
            "trader_name": "ChartMaster",
            "platform": "TradingView",
            "profile_url": "https://tradingview.com/u/chartmaster",
            "location": "Toronto, Canada",
            "language": "English",
            "strategy_type": "Technical Analysis",
            "twitter": "@chartmaster_tv",
            "youtube": "https://youtube.com/@chartmaster",
            "website": "https://chartmaster.io",
            "followers": 22000,
            "copiers": 0,
            "monthly_return": 8.0,
            "yearly_return": 96.0,
            "maximum_drawdown": 15.0,
            "risk_score": 5,
            "audience_strength": 10,
            "trading_consistency": 7,
            "communication_quality": 9,
            "crypto_knowledge": 6,
            "likelihood_to_join": 5,
            "brand_value": 8,
            "fit_score": 70,
            "creator_score": 68,
            "research_notes": "Large following on TV. More traditional TA. May not be DeFi-focused enough.",
            "tags": "Influencer, Technical Analysis, Large Audience",
            "status": "Rejected",
            "pipeline_stage": "Closed",
            "rejection_date": "2025-06-03T12:00:00Z",
            "date_added": "2025-05-15T09:00:00Z",
            "date_updated": "2025-06-03T12:00:00Z",
            "outreach_attempts": 1,
            "rejection_reason": "Strategy not aligned with DeFi copy-trading model. Focuses on traditional markets.",
        },
        {
            "trader_name": "ZuluWhale",
            "platform": "ZuluTrade",
            "profile_url": "https://zulutrade.com/trader/zw123456",
            "location": "Athens, Greece",
            "language": "Greek, English",
            "strategy_type": "Copy Trading",
            "followers": 5600,
            "copiers": 890,
            "assets_under_management": 4200000.0,
            "monthly_return": 5.5,
            "yearly_return": 66.0,
            "maximum_drawdown": 20.0,
            "risk_score": 8,
            "audience_strength": 6,
            "trading_consistency": 8,
            "communication_quality": 5,
            "crypto_knowledge": 4,
            "likelihood_to_join": 7,
            "brand_value": 5,
            "fit_score": 58,
            "creator_score": 55,
            "research_notes": "Experienced ZuluTrade provider. Some forex exposure. DeFi knowledge seems limited but willing to learn.",
            "tags": "Copy Trading, Potential",
            "status": "Negotiation",
            "pipeline_stage": "Contract",
            "date_added": "2025-05-10T07:00:00Z",
            "date_contacted": "2025-05-18T14:00:00Z",
            "date_replied": "2025-05-20T10:00:00Z",
            "date_interested": "2025-05-25T11:00:00Z",
            "date_negotiation_started": "2025-06-01T09:00:00Z",
            "last_contact_date": "2025-06-04T15:00:00Z",
            "date_updated": "2025-06-04T15:00:00Z",
            "outreach_attempts": 5,
        },
    ]
    for trader_data in samples:
        try:
            await create_trader(conn, trader_data)
        except Exception:
            continue
    logger.info("Sample traders seeded")


app = FastAPI(lifespan=lifespan)


def _deserialize_json_fields(trader):
    """Deserialize JSON string fields and escape HTML in user content."""
    if not trader:
        return trader
    # Escape HTML in raw_text to prevent XSS
    if 'raw_text' in trader and trader['raw_text']:
        trader['raw_text'] = escape_html(trader['raw_text'])
    return trader


# ===== Serve frontend =====

@app.get("/")
async def serve_index():
    return FileResponse(INDEX_PATH)


@app.get("/design-system.css")
async def serve_css():
    return FileResponse(CSS_PATH, media_type="text/css")


# ===== Traders API =====

@app.get("/api/traders")
async def list_traders(
    status: str = Query("all"),
    platform: str = Query("all"),
    min_fit: int = Query(0),
    search: str = Query(None),
    sort: str = Query("date_added DESC"),
):
    async with aiosqlite.connect(DB_PATH) as conn:
        conn.row_factory = aiosqlite.Row
        traders = await get_traders(
            conn=conn,
            status=status,
            platform=platform,
            min_fit=min_fit,
            search=search,
            sort=sort,
        )
    return [_deserialize_json_fields(trader) for trader in traders]


# ===== Stats API (must be before {trader_id} to avoid route conflict) =====

@app.get("/api/traders/stats")
async def get_stats():
    async with aiosqlite.connect(DB_PATH) as conn:
        conn.row_factory = aiosqlite.Row
        stats = await get_traders_stats(conn)
    return stats


@app.get("/api/traders/{trader_id}")
async def read_trader(trader_id: str):
    async with aiosqlite.connect(DB_PATH) as conn:
        conn.row_factory = aiosqlite.Row
        trader = await get_trader(conn, trader_id)
    if not trader:
        raise HTTPException(404, "Trader not found")
    return _deserialize_json_fields(trader)


@app.post("/api/traders")
async def add_trader(trader_data: dict):
    async with aiosqlite.connect(DB_PATH) as conn:
        conn.row_factory = aiosqlite.Row
        trader_id = await create_trader(conn, trader_data)
        trader = await get_trader(conn, trader_id)
    return trader


@app.put("/api/traders/{trader_id}")
async def edit_trader(trader_id: str, updates: dict):
    async with aiosqlite.connect(DB_PATH) as conn:
        conn.row_factory = aiosqlite.Row
        trader = await get_trader(conn, trader_id)
        if not trader:
            raise HTTPException(404, "Trader not found")
        updated = await update_trader(conn, trader_id, updates)
    return _deserialize_json_fields(updated)


@app.delete("/api/traders/{trader_id}")
async def remove_trader(trader_id: str):
    async with aiosqlite.connect(DB_PATH) as conn:
        deleted = await delete_trader(conn, trader_id)
    if not deleted:
        raise HTTPException(404, "Trader not found")
    return {"ok": True}


@app.delete("/api/traders")
async def clear_all_traders(confirm: str = ""):
    if confirm != "yes":
        raise HTTPException(400, "Send ?confirm=yes to clear all data")
    async with aiosqlite.connect(DB_PATH) as conn:
        await conn.execute("PRAGMA foreign_keys = ON")
        await conn.execute("DELETE FROM outreach_logs")
        await conn.execute("DELETE FROM traders")
        await conn.commit()
    return {"ok": True}


# ===== Analysis API =====

@app.post("/api/traders/{trader_id}/analyze")
async def analyze_trader(trader_id: str):
    async with aiosqlite.connect(DB_PATH) as conn:
        conn.row_factory = aiosqlite.Row
        trader = await get_trader(conn, trader_id)
        if not trader:
            raise HTTPException(404, "Trader not found")
        
        raw_text = trader.get('raw_text', '') or ''
        analysis = analyze_trader_keywords(raw_text)
        
        existing_notes = trader.get('research_notes', '') or ''
        keyword_line = f"\n\n[Auto Analysis {datetime.utcnow().strftime('%Y-%m-%d')}] Keywords: {', '.join(analysis['keywords'])}"
        await update_trader(conn, trader_id, {
            'research_notes': existing_notes + keyword_line
        })
    
    analysis["summary"] = escape_html(analysis["summary"])
    return analysis


# ===== Outreach API =====

@app.post("/api/traders/{trader_id}/outreach")
async def add_outreach(trader_id: str, outreach: dict):
    async with aiosqlite.connect(DB_PATH) as conn:
        conn.row_factory = aiosqlite.Row
        trader = await get_trader(conn, trader_id)
        if not trader:
            raise HTTPException(404, "Trader not found")
        method = outreach.get("method", "other")
        notes = outreach.get("notes", "")
        outreach_id = await log_outreach(
            conn=conn,
            trader_id=trader_id,
            method=method,
            notes=notes,
        )
        # Get updated trader with new outreach logs
        updated_trader = await get_trader(conn, trader_id)
    return {"id": outreach_id, "success": True, "trader": updated_trader}


# ===== Settings API =====

@app.get("/api/settings")
async def read_settings():
    async with aiosqlite.connect(DB_PATH) as conn:
        conn.row_factory = aiosqlite.Row
        settings = await get_all_settings(conn)
    return settings


@app.put("/api/settings")
async def write_settings(settings: dict):
    async with aiosqlite.connect(DB_PATH) as conn:
        conn.row_factory = aiosqlite.Row
        for key, value in settings.items():
            await set_setting(conn, key, str(value) if value is not None else "")
        updated = await get_all_settings(conn)
    return updated


# ===== Import/Export =====

@app.get("/api/export")
async def export_data():
    async with aiosqlite.connect(DB_PATH) as conn:
        conn.row_factory = aiosqlite.Row
        traders = await get_traders(conn)
        settings = await get_all_settings(conn)
    return {"traders": traders, "settings": settings}


@app.post("/api/import")
async def import_data(data: dict):
    if "traders" not in data or not isinstance(data["traders"], list):
        raise HTTPException(400, "Invalid format: expected {traders: [...]}")
    async with aiosqlite.connect(DB_PATH) as conn:
        count = await import_traders(conn, data["traders"])

    settings = data.get("settings")
    if settings and isinstance(settings, dict):
        async with aiosqlite.connect(DB_PATH) as conn:
            for key, value in settings.items():
                await set_setting(conn, key, str(value) if value is not None else "")

    return {"imported": count}


# ===== File Parse API =====

COLUMN_MAP = {
    'name': 'trader_name', 'trader': 'trader_name', 'trader_name': 'trader_name',
    'source': 'platform', 'platform': 'platform',
    'profile link': 'profile_url', 'url': 'profile_url', 'profile': 'profile_url', 'profile_url': 'profile_url',
    'linktree': '_social', 'social': '_social', 'twitter': 'twitter', 'telegram': 'telegram',
    'discord': 'discord', 'youtube': 'youtube', 'website': 'website',
    'message status': 'status', 'status': 'status', 'pipeline stage': 'pipeline_stage',
}

STATUS_MAP = {
    'messaged': 'Contacted', 'emailed': 'Contacted', 'contacted': 'Contacted',
    'replied': 'Replied', 'interested': 'Interested', 'meeting': 'Meeting Scheduled',
    'negotiation': 'Negotiation', 'onboarded': 'Onboarding', 'rejected': 'Rejected',
}


def _parse_social_links(text):
    """Extract twitter/telegram/discord from a URL or text string."""
    result = {}
    if not text:
        return result
    text = text.strip().lower()
    if 't.me/' in text or 'telegram' in text:
        result['telegram'] = text.split('t.me/')[-1].split('/')[0].split('?')[0] if 't.me/' in text else text
    elif 'twitter.com/' in text or 'x.com/' in text:
        result['twitter'] = text.split('twitter.com/')[-1].split('x.com/')[-1].split('/')[0].split('?')[0]
    elif 'discord' in text:
        result['discord'] = text
    elif 'youtube.com/' in text or 'youtu.be/' in text:
        result['youtube'] = text
    elif text.startswith('http'):
        result['website'] = text
    return result


def _map_row(headers, row):
    """Map a row of data using column headers."""
    trader = {}
    social_text = []
    for i, header in enumerate(headers):
        if i >= len(row):
            continue
        val = (row[i] or '').strip()
        if not val:
            continue
        key = header.strip().lower()
        field = COLUMN_MAP.get(key)
        if field == '_social':
            social_text.append(val)
        elif field == 'status':
            trader['status'] = STATUS_MAP.get(val.lower(), val)
        elif field:
            trader[field] = val
    for text in social_text:
        trader.update(_parse_social_links(text))
    return trader


@app.post("/api/parse-file")
async def parse_file(file: bytes = File(...), filename: str = Form("")):
    """Parse uploaded Excel or CSV file and return structured trader data."""
    rows = []
    headers = []
    try:
        if filename.endswith('.csv') or filename.endswith('.tsv'):
            text = file.decode('utf-8-sig')
            reader = csv.reader(io.StringIO(text))
            all_rows = list(reader)
            if not all_rows:
                raise HTTPException(400, "Empty file")
            headers = all_rows[0]
            for row in all_rows[1:]:
                if any(cell.strip() for cell in row):
                    rows.append(_map_row(headers, row))
        elif filename.endswith(('.xlsx', '.xls')):
            wb = load_workbook(io.BytesIO(file), read_only=True, data_only=True)
            ws = wb.active
            all_rows = [[str(cell or '') for cell in row] for row in ws.iter_rows()]
            wb.close()
            if not all_rows:
                raise HTTPException(400, "Empty file")
            headers = all_rows[0]
            for row in all_rows[1:]:
                if any(cell.strip() for cell in row):
                    rows.append(_map_row(headers, row))
        else:
            raise HTTPException(400, "Unsupported file type. Use .csv, .xlsx, or .xls")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"Parse error: {str(e)}")
    return {"rows": rows, "count": len(rows), "headers": headers}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=PORT)
